"""
工具函数模块 - 通用的辅助函数
"""
import re
import pandas as pd
import os
from openpyxl import load_workbook, Workbook

def parse_excel_range(range_str):
    """
    解析Excel范围字符串，返回工作表名和单元格范围
    """
    # 移除可能的单引号
    range_str = range_str.replace("'", "")
    
    if '!' in range_str:
        sheet_name, cell_range = range_str.split('!')
        return sheet_name.strip(), cell_range.strip()
    else:
        return None, range_str.strip()

def col_letter_to_index(letter):
    """
    将Excel列字母转换为数字索引
    """
    index = 0
    for i, c in enumerate(letter):
        index += (ord(c) - ord('A') + 1) * (26 ** (len(letter) - i - 1))
    return index - 1  # 转换为0-based索引

def write_results_to_excel(results_df, output_file, range_str):
    """
    将结果写入Excel的指定位置
    """
    try:
        sheet_name, cell_range = parse_excel_range(range_str)
        
        if ':' in cell_range:
            # 范围写入
            start_cell, end_cell = cell_range.split(':')
            
            # 解析起始单元格
            start_col_letter = re.findall('[A-Z]+', start_cell)[0]
            start_row = int(re.findall('[0-9]+', start_cell)[0])
            
            start_col = col_letter_to_index(start_col_letter) + 1  # 转换为1-based索引
            
            # 使用openpyxl直接写入Excel
            # 加载或创建Excel文件
            if os.path.exists(output_file):
                book = load_workbook(output_file)
            else:
                book = Workbook()
            
            # 选择或创建工作表
            if sheet_name in book.sheetnames:
                sheet = book[sheet_name]
            else:
                sheet = book.create_sheet(sheet_name)
            
            # 写入数据
            for i, row in results_df.iterrows():
                for j, value in enumerate(row):
                    sheet.cell(row=start_row + i, column=start_col + j, value=value)
            
            # 保存文件
            book.save(output_file)
            print(f"结果已写入 {output_file} 的 {range_str}")
            
        else:
            # 单个单元格写入（只写入第一个值）
            if os.path.exists(output_file):
                book = load_workbook(output_file)
            else:
                book = Workbook()
            
            if sheet_name in book.sheetnames:
                sheet = book[sheet_name]
            else:
                sheet = book.create_sheet(sheet_name)
            
            # 解析单元格
            col_letter = re.findall('[A-Z]+', cell_range)[0]
            row_num = int(re.findall('[0-9]+', cell_range)[0])
            
            col_num = col_letter_to_index(col_letter) + 1  # 转换为1-based索引
            
            # 写入第一个值
            if len(results_df) > 0 and len(results_df.columns) > 0:
                sheet.cell(row=row_num, column=col_num, value=results_df.iloc[0, 0])
            
            book.save(output_file)
            print(f"结果已写入 {output_file} 的 {range_str}")
            
    except Exception as e:
        print(f"写入结果到Excel时出错: {e}")
        raise

def validate_data(data):
    """
    验证读取的数据是否有效
    """
    print("=== 数据验证 ===")
    print(f"beta_min 样本: {data['beta_min'][:5]}")
    print(f"beta_max 样本: {data['beta_max'][:5]}")
    print(f"alpha_min 样本: {data['alpha_min'][:5]}")
    print(f"alpha_max 样本: {data['alpha_max'][:5]}")
    print(f"SKU beta 样本: {data['beta'][:5]}")
    print(f"SKU alpha 样本: {data['alpha'][:5]}")

    '''
    # 验证约束可行性
    for k in range(min(5, data['num_of_Batches'])):
        valid_skus = [i for i in range(data['num_of_SKUs']) 
                    if (data['beta'][i] >= data['beta_min'][k] and 
                        data['beta'][i] < data['beta_max'][k] and
                        data['alpha'][i] < data['alpha_max'][k])]
        print(f"批次 {k}: 符合条件的SKU数量 = {len(valid_skus)}")
    '''

def write_df_to_sheet(
    output_file: str,
    sheet_name: str,
    df: pd.DataFrame,
    start_row: int = 1,
    start_col: int = 1,
    clear_sheet: bool = False
):
    """将DataFrame写入指定Excel的指定sheet，从(start_row, start_col)开始写。
    - 默认写表头到 start_row 行
    - clear_sheet=True 时，会清空该sheet已有内容（谨慎使用）
    """
    if os.path.exists(output_file):
        book = load_workbook(output_file)
    else:
        book = Workbook()

    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
    else:
        sheet = book.create_sheet(sheet_name)

    if clear_sheet:
        sheet.delete_rows(1, sheet.max_row)

    # 写表头
    for j, col in enumerate(df.columns, start=start_col):
        sheet.cell(row=start_row, column=j, value=col)

    # 写数据
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=start_col):
            val = row[col]
            # numpy 标量转 Python 标量，避免 openpyxl 类型问题
            try:
                import numpy as np
                if isinstance(val, (np.generic,)):
                    val = val.item()
            except Exception:
                pass
            sheet.cell(row=i, column=j, value=val)

    book.save(output_file)

# bld_check.py
# 输入粗检查（只使用 SKU 数 I 与车辆数 V）

from dataclasses import dataclass
from typing import Tuple, Optional


@dataclass
class PrecheckResult:
    ok: bool
    message: str


def precheck_counts(num_skus: int, num_vehicles: int):
    # ==========================
    # 1) 基础合法性检查
    # ==========================
    if num_skus <= 0 or num_vehicles <= 0:
        return PrecheckResult(
            ok=False,
            message="[error]SKU数量和车次数量必须大于 0。"
        )

    # ==========================
    # 2) 经验规则：SKU 数不能过少
    #    I >= 30*V - 30
    # ==========================
    expected = 30 * num_vehicles
    lb = max(1, expected - 30)

    if num_skus < lb:
        return PrecheckResult(
            ok=False,
            message=(
                f"[error]订单参数预检查未通过：SKU 数量过少。"
                f"当前 SKU数量为：{num_skus}, 车次数量为：{num_vehicles}；"
                f"建议 SKU数量 ≥ {lb}。"
            )
        )

    return PrecheckResult(
        ok=True,
        message="[success]订单参数预检查通过。"
    )


'''
def precheck_from_data(data) -> PrecheckResult:
    I = int(data.num_of_SKUs)
    V = int(data.num_of_Vehicles)

    # 基本合法性
    if I <= 0 or V <= 0:
        return PrecheckResult(False, f"[error]预检查失败：I={I}, V={V} 必须都 > 0。")

    # (可选) 若你要求每车至少 1 个SKU
    if I < V:
        return PrecheckResult(False, f"[error]预检查失败：I={I} < V={V}，无法保证每车非空。")

    # 关键：重量下界硬检查（建议加入）
    total_w = float(sum(data.w))
    need_w = float(V) * float(data.W_v_min)
    if total_w < need_w:
        return PrecheckResult(
            False,
            f"[error]预检查失败：SKU总湿重={total_w:.3f} < V*W_v_min={need_w:.3f} "
            f"(V={V}, W_v_min={data.W_v_min})，必不可行。"
        )

    return PrecheckResult(True, f"[pass]预检查通过：I={I}, V={V}，总湿重={total_w:.3f}。")
'''
