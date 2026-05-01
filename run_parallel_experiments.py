import csv
import shutil
import subprocess
import sys
import time
from pathlib import Path

from openpyxl import load_workbook

# ==============================
# 并行批跑配置
# i -> Params!B1 -> num_of_SKUs
# v -> Params!B2 -> num_of_Vehicles
# t -> 文件命名标签，也是 warmstart 注入后整体 MIP 主求解阶段的时间上限
# ==============================
BASE_DIR = Path(__file__).resolve().parent
BASE_INPUT = BASE_DIR / "input.xlsx"
MAIN_SCRIPT = BASE_DIR / "bld_main.py"
PYTHON_EXE = sys.executable

SCENARIOS = [
    # (200, 2),
    # (300, 2),
    # (200, 3),
    # (300, 3),
    (200, 4),
    (200, 5),
    (300, 10),
]

# 逐车求解得到初始解并注入 SCIP 后，整体 MIP 主求解阶段的强制运行时间。
MAIN_TIME_LIMIT_SEC = 300

# 预留接口：控制逐车 warmstart 阶段的总预算，对应 Excel Params!B12。
# 默认 None 表示不覆盖基础 input.xlsx 里的 B12。
# 求解时会按“第 k 辆车预算 = (总预算 - 已用时间) / 剩余车辆数”动态分配。
# 例如想把逐车 warmstart 阶段总预算设为 300 秒，可改为：WARMSTART_TIME_LIMIT_SEC = 300
WARMSTART_TIME_LIMIT_SEC = 300

# 预留接口：给 warmstart 阶段每个单车 MIP 再加一个附加上限。
# 默认 None 表示只使用 WARMSTART_TIME_LIMIT_SEC / Params!B12 的动态平均分配策略。
# 如需让任何单车最多 300 秒，可改为：WARMSTART_VEHICLE_TIME_LIMIT_SEC = 300
# 实际单车时限 = min(动态平均预算, WARMSTART_VEHICLE_TIME_LIMIT_SEC)。
WARMSTART_VEHICLE_TIME_LIMIT_SEC = None

# 是否禁用 warmstart 做冷启动对照实验。
DISABLE_WARMSTART = False

# 并发数。SCIP/MIP 较吃 CPU 和内存，建议从 1 或 2 开始。
MAX_CONCURRENT = min(4, len(SCENARIOS))
POLL_INTERVAL_SEC = 2.0

ROOT = BASE_DIR / "parallel_runs"
INPUT_DIR = ROOT / "inputs"
OUTPUT_DIR = ROOT / "outputs"
LOG_DIR = ROOT / "logs"
CSV_DIR = ROOT / "csv"
REPORT_DIR = ROOT / "reports"

for directory in [INPUT_DIR, OUTPUT_DIR, LOG_DIR, CSV_DIR, REPORT_DIR]:
    directory.mkdir(parents=True, exist_ok=True)


def build_case_tag(i_value: int, v_value: int, timelimit_sec: int) -> str:
    return f"i{i_value}v{v_value}t{timelimit_sec}"


def prepare_input_file(base_input: Path, target_input: Path, i_value: int, v_value: int):
    shutil.copy2(base_input, target_input)

    wb = load_workbook(target_input)
    ws = wb["Params"]
    ws["B1"] = int(i_value)
    ws["B2"] = int(v_value)

    # 预留接口：只在显式设置时覆盖逐车 warmstart 阶段使用的 Params!B12。
    if WARMSTART_TIME_LIMIT_SEC is not None:
        ws["B12"] = float(WARMSTART_TIME_LIMIT_SEC)

    wb.save(target_input)
    wb.close()


def expected_progress_csv(output_file: Path) -> Path:
    return output_file.parent / f"{output_file.stem}_scip_progress" / "scip_progress_data.csv"


def expected_scip_raw_log(output_file: Path) -> Path:
    return output_file.parent / f"{output_file.stem}_scip_progress" / "scip_raw.log"


def launch_one_case(i_value: int, v_value: int) -> dict:
    case_tag = build_case_tag(i_value, v_value, MAIN_TIME_LIMIT_SEC)

    input_file = INPUT_DIR / f"{case_tag}_input.xlsx"
    output_file = OUTPUT_DIR / f"{case_tag}.xlsx"
    log_file = LOG_DIR / f"{case_tag}.log"
    csv_target = CSV_DIR / f"{case_tag}.csv"
    raw_log_target = LOG_DIR / f"{case_tag}_scip_raw.log"

    prepare_input_file(BASE_INPUT, input_file, i_value, v_value)

    cmd = [
        PYTHON_EXE,
        str(MAIN_SCRIPT),
        "--mode", "solve",
        "--input", str(input_file),
        "--output", str(output_file),
        "--main-time-limit", str(MAIN_TIME_LIMIT_SEC),
    ]
    if WARMSTART_VEHICLE_TIME_LIMIT_SEC is not None and float(WARMSTART_VEHICLE_TIME_LIMIT_SEC) > 0:
        cmd.extend(["--vehicle-time-limit", str(float(WARMSTART_VEHICLE_TIME_LIMIT_SEC))])
    if DISABLE_WARMSTART:
        cmd.append("--disable-warmstart")

    for stale_path in [log_file, csv_target, raw_log_target]:
        try:
            if stale_path.exists():
                stale_path.unlink()
        except Exception:
            pass

    log_f = open(log_file, "w", encoding="utf-8")
    log_f.write(f"[batch] case_tag = {case_tag}\n")
    log_f.write(f"[batch] command = {' '.join(cmd)}\n")
    log_f.write(f"[batch] MAIN_TIME_LIMIT_SEC = {MAIN_TIME_LIMIT_SEC}\n")
    log_f.write(f"[batch] WARMSTART_TIME_LIMIT_SEC = {WARMSTART_TIME_LIMIT_SEC}\n")
    log_f.write("[batch] warmstart 单车动态预算 = (总预算 - 已用时间) / 剩余车辆数\n")
    log_f.write(f"[batch] WARMSTART_VEHICLE_TIME_LIMIT_SEC = {WARMSTART_VEHICLE_TIME_LIMIT_SEC}\n")
    log_f.write(f"[batch] DISABLE_WARMSTART = {DISABLE_WARMSTART}\n\n")
    log_f.flush()

    proc = subprocess.Popen(
        cmd,
        stdout=log_f,
        stderr=subprocess.STDOUT,
        cwd=str(BASE_DIR),
    )

    return {
        "case_tag": case_tag,
        "i_value": i_value,
        "v_value": v_value,
        "proc": proc,
        "log_f": log_f,
        "input_file": input_file,
        "output_file": output_file,
        "log_file": log_file,
        "csv_source": expected_progress_csv(output_file),
        "csv_target": csv_target,
        "raw_log_source": expected_scip_raw_log(output_file),
        "raw_log_target": raw_log_target,
        "return_code": None,
        "finished": False,
    }


def finalize_job(job: dict):
    if job["finished"]:
        return

    ret = job["proc"].wait()
    job["log_f"].close()
    job["return_code"] = ret
    job["finished"] = True

    if ret == 0:
        csv_source = job["csv_source"]
        csv_target = job["csv_target"]
        raw_log_source = job["raw_log_source"]
        raw_log_target = job["raw_log_target"]

        if csv_source.exists():
            shutil.copy2(csv_source, csv_target)
            print(f"[done] {job['case_tag']} 运行完成，CSV：{csv_target}")
        else:
            print(f"[warn] {job['case_tag']} 运行完成，但未找到 CSV：{csv_source}")

        if raw_log_source.exists():
            shutil.copy2(raw_log_source, raw_log_target)
            print(f"[done] {job['case_tag']} SCIP 原生日志：{raw_log_target}")
    else:
        print(f"[failed] {job['case_tag']} 运行失败，返回码={ret}，请查看日志：{job['log_file']}")


def write_summary(jobs: list[dict]):
    summary_file = REPORT_DIR / "batch_run_summary.csv"
    with summary_file.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "case_tag", "i_value", "v_value", "return_code",
                "main_time_limit_sec", "warmstart_time_limit_sec", "warmstart_vehicle_time_limit_sec",
                "input_file", "output_file", "log_file", "csv_target", "raw_log_target",
            ],
        )
        writer.writeheader()
        for job in jobs:
            writer.writerow({
                "case_tag": job["case_tag"],
                "i_value": job["i_value"],
                "v_value": job["v_value"],
                "return_code": job["return_code"],
                "main_time_limit_sec": MAIN_TIME_LIMIT_SEC,
                "warmstart_time_limit_sec": WARMSTART_TIME_LIMIT_SEC,
                "warmstart_vehicle_time_limit_sec": WARMSTART_VEHICLE_TIME_LIMIT_SEC,
                "input_file": str(job["input_file"]),
                "output_file": str(job["output_file"]),
                "log_file": str(job["log_file"]),
                "csv_target": str(job["csv_target"]),
                "raw_log_target": str(job["raw_log_target"]),
            })
    print(f"[summary] 汇总文件：{summary_file}")


def main():
    if not BASE_INPUT.exists():
        raise FileNotFoundError(f"未找到基础输入文件：{BASE_INPUT}")
    if not MAIN_SCRIPT.exists():
        raise FileNotFoundError(f"未找到主程序文件：{MAIN_SCRIPT}")

    pending = list(SCENARIOS)
    running = []
    finished = []

    print(f"[info] 共 {len(SCENARIOS)} 个场景，最大并发数 = {MAX_CONCURRENT}")
    print(f"[info] 主 MIP 强制时间 = {MAIN_TIME_LIMIT_SEC} 秒")
    print(f"[info] 逐车 warmstart 阶段总预算 B12 覆盖 = {WARMSTART_TIME_LIMIT_SEC}")
    print("[info] warmstart 单车动态预算策略：第 k 辆车预算 = (总预算 - 已用时间) / 剩余车辆数")
    print(f"[info] warmstart 单车附加上限接口 = {WARMSTART_VEHICLE_TIME_LIMIT_SEC}")
    print(f"[info] warmstart = {'关闭' if DISABLE_WARMSTART else '开启'}")

    while pending or running:
        while pending and len(running) < MAX_CONCURRENT:
            i_value, v_value = pending.pop(0)
            job = launch_one_case(i_value, v_value)
            running.append(job)
            print(f"[started] {job['case_tag']}")

        still_running = []
        for job in running:
            ret = job["proc"].poll()
            if ret is None:
                still_running.append(job)
            else:
                finalize_job(job)
                finished.append(job)

        running = still_running

        if running:
            time.sleep(POLL_INTERVAL_SEC)

    success_count = sum(1 for job in finished if job["return_code"] == 0)
    fail_count = len(finished) - success_count
    print("\n全部场景已结束。")
    print(f"[summary] 成功：{success_count}，失败：{fail_count}")
    print(f"[summary] 输入目录：{INPUT_DIR}")
    print(f"[summary] 输出目录：{OUTPUT_DIR}")
    print(f"[summary] 日志目录：{LOG_DIR}")
    print(f"[summary] CSV 目录：{CSV_DIR}")
    write_summary(finished)

    if fail_count > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
